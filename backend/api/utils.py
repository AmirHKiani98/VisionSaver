from collections import defaultdict
import pandas as pd
from record.models import Record
from ai.models import AutoDetection, DetectionLines
from record.models import RecordLog
from datetime import timedelta
import requests
import os
import re
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def get_ip_from_rtsp(rtsp_link):
    return re.findall(r"rtsp://(\d+\.\d+\.\d+\.\d+)", rtsp_link)[0]
def get_movement_index(movement):
    if "through" in movement.lower():
        return 0
    elif "left" in movement.lower():
        return 1
    elif "right" in movement.lower():
        return 2
    return -1 

def consistent_direction(movement):
    if "through" in movement.lower():
        return "through"
    elif "left" in movement.lower():
        return "left"
    elif "right" in movement.lower():
        return "right"
    return "unknown"


def all_complete_results_from_record(record_id):
    """
    Get the excel pandas file from data
    """
    result = pd.DataFrame({"time": [], "count": [], "turn_movement": [], "type": [], "approach_direction": []})
    record = Record.objects.filter(id=record_id).first()
    if record is None:
        return result
    approach_direction = record.direction or f"N/A, Record {record_id}"
    start_time = record.start_time
    iss_detections, _ = get_iss_detections_pandas(record_id)
    manual_counts, _ = get_counter_manual_results(record_id)
    auto_record = AutoDetection.objects.filter(record=record).first()
    if auto_record:
        auto_counts, _ = get_counter_auto_detection_results(record_id, auto_record.version, auto_record.divide_time)
    else:
        auto_counts = {}
    if iss_detections is not None and not iss_detections.empty:
        iss_detections["count"] = 1
        iss_detections["turn_movement"] = iss_detections["direction"].apply(consistent_direction)
        iss_detections["type"] = "iss"
        iss_detections["approach_direction"] = approach_direction
        iss_detections = iss_detections[["time", "count", "turn_movement", "type", "approach_direction"]]
        iss_detections["time"] = pd.to_datetime(iss_detections["time"]).dt.tz_convert("UTC")
        # Append to result
        result = pd.concat([result, iss_detections], ignore_index=True)
    if manual_counts:
        for movement, data in manual_counts.items():
            for time, count in data.items():
                result = pd.concat([result, pd.DataFrame({
                    "time": [start_time + timedelta(seconds=time)],
                    "count": [count],
                    "turn_movement": [consistent_direction(movement)],
                    "type": ["manual"],
                    "approach_direction": [approach_direction]
                })], ignore_index=True)

    if auto_counts:
        for movement, data in auto_counts.items():
            for time, count_data in data.items():
                result = pd.concat([result, pd.DataFrame({
                    "time": [start_time + timedelta(seconds=time)],
                    "count": [count_data[0]],
                    "turn_movement": [consistent_direction(movement)],
                    "type": ["auto"],
                    "approach_direction": [approach_direction]
                })], ignore_index=True)
    result["interval"] = result["time"].apply(bucketize_time)

    return result


def get_complete_df_from_multiple_records(record_ids):
    final_result = pd.DataFrame({"time": [], "count": [], "turn_movement": [], "type": [], "approach_direction": []})
    for record_id in record_ids:
        record_result = all_complete_results_from_record(record_id)
        final_result = pd.concat([final_result, record_result], ignore_index=True)

    return final_result

def complete_categorized_df_to_wb(df, title=""):
    
    wb = openpyxl.Workbook()
    try:
        default = wb.active
        wb.remove(default)
    except Exception:
        pass
    # First row: empty | direction 1 | direction 2 | ...
    # Seconds row: Time | |Movement Type 1| |Movement Type 2| |Movement Type 3|| ...|...| ... || ...    
    # Third row onwards: |Time 1 | Count | Count | Count | ... | ... | ... || ... | ... |
    columns_should_exist = ["interval", "count", "turn_movement", "type", "approach_direction"]
    for column in columns_should_exist:
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found in DataFrame.")
    
    df = df.sort_values(["interval"])
    grouped = df.groupby(["interval", "approach_direction", "turn_movement", "type"]).agg({'count': 'sum'})
    grouped = grouped.reset_index()
    grouped = grouped.sort_values(["interval", "approach_direction", "turn_movement","type"])
    
    comparison_df = pd.DataFrame({})
    # Prepare comparison_df
    manual_df = grouped[grouped['type'] == 'manual'].copy()
    auto_df = grouped[grouped['type'] == 'auto'].copy()
    iss_df = grouped[grouped['type'] == 'iss'].copy()

    for index, row in auto_df.iterrows():
        interval = row['interval']
        approach_direction = row['approach_direction']
        turn_movement = row['turn_movement']
        auto_count = row['count']
        
        manual_count = manual_df[
            (manual_df['interval'] == interval) & 
            (manual_df['approach_direction'] == approach_direction) & 
            (manual_df['turn_movement'] == turn_movement)
        ]
        if not manual_count.empty:
            manual_count = manual_count.iloc[0]['count']
            # avoid division by zero
            if manual_count != 0:
                difference_percentage = abs(auto_count - manual_count) / manual_count
            else:
                difference_percentage = None
            comparison_df = pd.concat([comparison_df, pd.DataFrame({
                'interval': [interval],
                'approach_direction': [approach_direction],
                'turn_movement': [turn_movement],
                'count': [difference_percentage],
                'type': ['auto_vs_manual']
            })], ignore_index=True)

    for index, row in iss_df.iterrows():
        interval = row['interval']
        approach_direction = row['approach_direction']
        turn_movement = row['turn_movement']
        iss_count = row['count']
        
        manual_count = manual_df[
            (manual_df['interval'] == interval) & 
            (manual_df['approach_direction'] == approach_direction) & 
            (manual_df['turn_movement'] == turn_movement)
        ]
        if not manual_count.empty:
            manual_count = manual_count.iloc[0]['count']
            if manual_count != 0:
                difference_percentage = abs(iss_count - manual_count) / manual_count
            else:
                difference_percentage = None
            comparison_df = pd.concat([comparison_df, pd.DataFrame({
                'interval': [interval],
                'approach_direction': [approach_direction],
                'turn_movement': [turn_movement],
                'count': [difference_percentage],
                'type': ['iss_vs_manual']
            })], ignore_index=True)
    
    # build mapping of interval -> target excel row
    unique_intervals = sorted(grouped["interval"].unique().tolist())
    time_interval_row_mapping = {interval: idx + 3 for idx, interval in enumerate(unique_intervals)}

    # Styling helpers (match other exporters)
    from openpyxl.styles import PatternFill, Border, Side, Font
    from openpyxl.formatting.rule import ColorScaleRule
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    time_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fills = [
        PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    ]
    bold_font = Font(bold=True)

    # Build Time rows once for all sheets (we'll reuse the same intervals)
    # For each 'type' create a sheet with grouped layout
    grouped_by_type = grouped.groupby("type")
    for type_idx, (type_name, type_group) in enumerate(grouped_by_type):
        # Create sheet for each type
        ws = wb.create_sheet(title=f"{type_name} {title}"[:31])
        # top-left empty cell
        ws.cell(row=1, column=1, value="")
        ws.cell(row=2, column=1, value="Time")
        # fill time column values
        for interval, row_idx in time_interval_row_mapping.items():
            # ensure interval is datetime-like and tz-naive for Excel
            try:
                ts = pd.to_datetime(interval)
                if hasattr(ts, "tzinfo") and ts.tzinfo is not None:
                    ts = ts.tz_convert("UTC").tz_localize(None)
                ws.cell(row=row_idx, column=1, value=ts.strftime("%I:%M %p"))
            except Exception:
                ws.cell(row=row_idx, column=1, value=str(interval))

        # build columns grouped by approach_direction
        col_idx = 2
        type_df_grouped_by_approach_direction = type_group.groupby("approach_direction")
        # Collect mapping for styling (approach_direction -> start_col, end_col, num_movements)
        mapping = []
        for approach_direction, approach_group in type_df_grouped_by_approach_direction:
            movement_types = list(approach_group["turn_movement"].unique())
            num_movements = len(movement_types)
            if num_movements == 0:
                continue
            start_col = col_idx
            end_col = col_idx + num_movements - 1
            # Merge cells for approach direction header (row1)
            start_col_letter = get_column_letter(start_col)
            end_col_letter = get_column_letter(end_col)
            merge_range = f"{start_col_letter}1:{end_col_letter}1"
            try:
                ws.merge_cells(merge_range)
            except Exception:
                pass
            ws[f"{start_col_letter}1"] = approach_direction
            ws[f"{start_col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")
            # write movement types on row2 and fill in counts
            for movement in movement_types:
                cell = ws.cell(row=2, column=col_idx, value=movement)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Fill counts for this movement
                movement_data = approach_group[approach_group["turn_movement"] == movement]
                for _, row in movement_data.iterrows():
                    interval = row["interval"]
                    count = row["count"]
                    row_idx = time_interval_row_mapping[interval]
                    ws.cell(row=row_idx, column=col_idx, value=count)
                col_idx += 1
            mapping.append((approach_direction, start_col, end_col, num_movements))
        # Apply styling: fills, borders, column widths
        # time column styling
        time_col_idx = 1
        ws.cell(row=1, column=time_col_idx).fill = time_fill
        ws.cell(row=2, column=time_col_idx).fill = time_fill
        ws.column_dimensions[get_column_letter(time_col_idx)].width = 20

        # apply fills to merged headers and second row and set column widths
        for idx, (approach_direction, start_idx, end_idx, num_cols) in enumerate(mapping):
            fill = fills[idx % len(fills)]
            for col in range(start_idx, end_idx + 1):
                # header row (row1) cells inside merged area
                hcell = ws.cell(row=1, column=col)
                hcell.fill = fill
                hcell.border = border
                hcell.font = bold_font
                # second row header
                h2 = ws.cell(row=2, column=col)
                h2.fill = fill
                h2.border = border
                # set width per column (spread 90 units)
                ws.column_dimensions[get_column_letter(col)].width = max(8, 90.0 / max(1, num_cols))
        # set border/time fill for time header cells
        ws.cell(row=1, column=1).border = border
        ws.cell(row=2, column=1).border = border

        # Data rows: write row cells and apply fills/borders
        for interval, row_idx in time_interval_row_mapping.items():
            for col in range(1, col_idx):
                value = ws.cell(row=row_idx, column=col).value
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col == time_col_idx:
                    cell.fill = time_fill
                else:
                    # find fill based on mapping block
                    for idx, (_, start_idx, end_idx, _) in enumerate(mapping):
                        if start_idx <= col <= end_idx:
                            cell.fill = fills[idx % len(fills)]
                            break

        # Outer border for each approach_direction block (header + data rows)
        last_data_row = 2 + len(time_interval_row_mapping)
        for _, start_idx, end_idx, _ in mapping:
            for r in range(1, last_data_row + 1):
                for c in range(start_idx, end_idx + 1):
                    ws.cell(row=r, column=c).border = border

    # --- Add comparison sheets (Auto and ISS) with percentage and color scale ---
    def _write_comparison_sheet(sheet_title, comp_df):
        ws_cmp = wb.create_sheet(title=sheet_title[:31])
        if comp_df.empty:
            ws_cmp.append(["No comparison data"])
            return ws_cmp
        # columns: interval, approach_direction, turn_movement, count (percentage)
        headers = ["Time", "Approach Direction", "Turn Movement", "Difference"]
        for j, h in enumerate(headers, start=1):
            cell = ws_cmp.cell(row=1, column=j, value=h)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = bold_font
            cell.border = border
            if j == 1:
                ws_cmp.column_dimensions[get_column_letter(j)].width = 20
            else:
                ws_cmp.column_dimensions[get_column_letter(j)].width = 18

        # write rows
        rows = []
        for _, r in comp_df.iterrows():
            interval = r["interval"]
            try:
                ts = pd.to_datetime(interval)
                if hasattr(ts, "tzinfo") and ts.tzinfo is not None:
                    ts = ts.tz_convert("UTC").tz_localize(None)
                tval = ts.strftime("%Y-%m-%d %I:%M %p")
            except Exception:
                tval = str(interval)
            diff = r["count"] if pd.notna(r["count"]) else None
            rows.append([tval, r.get("approach_direction", ""), r.get("turn_movement", ""), diff])
        for i, row in enumerate(rows, start=2):
            for j, v in enumerate(row, start=1):
                cell = ws_cmp.cell(row=i, column=j, value=v)
                cell.border = border
                if j == 4:
                    # difference column: set percent format when value present
                    if v is not None:
                        try:
                            cell.value = float(v)
                            cell.number_format = "0.00%"
                        except Exception:
                            pass

        # apply color scale to Difference column (col 4)
        last_row = 1 + len(rows)
        diff_col_letter = get_column_letter(4)
        rule = ColorScaleRule(
            start_type='num', start_value=0, start_color="63BE7B",    # green
            mid_type='num', mid_value=0.5, mid_color="FFEB84",         # yellow
            end_type='num', end_value=1.0, end_color="FF0000"          # red at 100%
        )
        try:
            ws_cmp.conditional_formatting.add(f"{diff_col_letter}2:{diff_col_letter}{last_row}", rule)
        except Exception:
            pass
        return ws_cmp

    auto_comp = comparison_df[comparison_df["type"] == "auto_vs_manual"].copy() if not comparison_df.empty else pd.DataFrame()
    iss_comp = comparison_df[comparison_df["type"] == "iss_vs_manual"].copy() if not comparison_df.empty else pd.DataFrame()

    _write_comparison_sheet("Auto vs Manual Comparison", auto_comp)
    _write_comparison_sheet("ISS vs Manual Comparison", iss_comp)

    return wb

def bucketize_time(time):
    minute = (time.minute // 15) * 15
    return time.replace(minute=minute, second=0, microsecond=0)

def get_complete_results_for_multiple_record(record_ids):
    final_result = pd.DataFrame({"time": [], "count": [], "turn_movement": [], "type": [], "approach_direction": []})
    for record_id in record_ids:
        record_result = all_complete_results_from_record(record_id)
        final_result = pd.concat([final_result, record_result], ignore_index=True)
    
    # Time should be converted into Central Timezone
    if not pd.api.types.is_datetime64tz_dtype(final_result["time"]):
        final_result["time"] = pd.to_datetime(final_result["time"]).dt.tz_localize("UTC")
    final_result["time"] = final_result["time"].dt.tz_convert("America/Chicago")
    
    return final_result



def copy_sheet_to_workbook(src_ws_or_wb, dst_wb, dst_title=None):
    """
    Copy a worksheet or all worksheets from a workbook into dst_wb.
    - src_ws_or_wb: openpyxl.worksheet.worksheet.Worksheet OR openpyxl.workbook.workbook.Workbook
    - dst_title: for a single sheet copy, used as sheet title; for workbook input, used as prefix.
    Returns list of created worksheets (or single worksheet wrapped in list).
    """
    from copy import copy as _copy
    import openpyxl

    def _copy_one(src_ws, dst_wb, title):
        title = (title or src_ws.title)[:31]
        dst_ws = dst_wb.create_sheet(title=title)
        # copy column widths
        for col, dim in src_ws.column_dimensions.items():
            try:
                dst_ws.column_dimensions[col].width = dim.width
            except Exception:
                pass
        # copy row heights
        for idx, rd in src_ws.row_dimensions.items():
            try:
                dst_ws.row_dimensions[idx].height = rd.height
            except Exception:
                pass
        # copy merged cells
        for merged in list(src_ws.merged_cells.ranges):
            try:
                dst_ws.merge_cells(str(merged))
            except Exception:
                pass
        # copy panes/freeze
        try:
            dst_ws.sheet_view = src_ws.sheet_view
        except Exception:
            pass
        # copy row/column visibility, outline levels
        try:
            for idx, rd in src_ws.row_dimensions.items():
                try:
                    dst_ws.row_dimensions[idx].hidden = rd.hidden
                    dst_ws.row_dimensions[idx].outlineLevel = rd.outlineLevel
                except Exception:
                    pass
            for col, cd in src_ws.column_dimensions.items():
                try:
                    dst_ws.column_dimensions[col].hidden = cd.hidden
                    dst_ws.column_dimensions[col].outlineLevel = cd.outlineLevel
                except Exception:
                    pass
        except Exception:
            pass
        # copy cells
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                try:
                    if cell.has_style:
                        try: new_cell.font = _copy(cell.font)
                        except Exception: pass
                        try: new_cell.border = _copy(cell.border)
                        except Exception: pass
                        try: new_cell.fill = _copy(cell.fill)
                        except Exception: pass
                        try: new_cell.alignment = _copy(cell.alignment)
                        except Exception: pass
                        try: new_cell.protection = _copy(cell.protection)
                        except Exception: pass
                        try: new_cell.number_format = cell.number_format
                        except Exception: pass
                    try:
                        if cell.comment is not None:
                            new_cell.comment = _copy(cell.comment)
                    except Exception:
                        pass
                    try:
                        if cell.hyperlink is not None:
                            new_cell.hyperlink = _copy(cell.hyperlink)
                    except Exception:
                        pass
                    try:
                        if hasattr(cell, "_style"):
                            new_cell._style = _copy(cell._style)
                    except Exception:
                        pass
                except Exception:
                    pass
        return dst_ws

    created = []
    # detect workbook vs worksheet
    try:
        is_wb = hasattr(src_ws_or_wb, "worksheets")
    except Exception:
        is_wb = False

    if is_wb:
        prefix = (dst_title or "").strip()
        for ws in src_ws_or_wb.worksheets:
            # build title; if prefix provided, use prefix + " - " + sheetname
            if prefix:
                candidate = f"{prefix} - {ws.title}"
            else:
                candidate = ws.title
            created.append(_copy_one(ws, dst_wb, candidate))
    else:
        # single worksheet
        created.append(_copy_one(src_ws_or_wb, dst_wb, dst_title))
    return created

def get_auto_detection_results_from_df_zones(auto_df, min_time=0, max_time=0, lines_map_length=None):
    results = defaultdict(dict)
    if max_time == 0:
        max_time = auto_df['time'].max()
    
    print("Max time set to", max_time)
    auto_df = auto_df[(auto_df['time'] >= min_time) & (auto_df['time'] <= max_time)]
    THRESHOLD = 0.5
    auto_df = auto_df[auto_df["confidence"] >= THRESHOLD]
    auto_df = auto_df.sort_values(["time", "track_id"])
    grouped = auto_df.groupby('track_id')['time'].agg(['min', 'max'])
    grouped['duration'] = grouped["max"] - grouped["min"]
    grouped = grouped.reset_index()


    grouped = grouped[grouped["duration"] > 2]
    track_ids_of_interest = grouped["track_id"].unique()

    filtered_df = auto_df[auto_df['track_id'].isin(track_ids_of_interest)]
    groups = filtered_df.groupby("track_id")
    total = 0
    for _, group in groups:
        detected = group[group["in_area"]]
        unique_line_indexes = detected["line_index"].unique().tolist()
        for line_index in unique_line_indexes:
            zone_detected = detected[detected["line_index"] == line_index]["zone_index"].unique()
            if len(zone_detected) == lines_map_length.get(line_index, 0) and list(zone_detected) == sorted(zone_detected):
                # Check the sort of the nu_zone_detected too
                # e.g. [2, 1, 3] is not accetable. Only [1,2,3]
                time = detected[detected["line_index"] == line_index]["time"].max()
                if time not in results[line_index]:
                    results[line_index][time] = [0, [], [], []]
                results[line_index][time][0] += 1
                total += 1
                results[line_index][time][1].append(int(group["track_id"].iloc[0]))
                results[line_index][time][2].append(group["cls_id"].iloc[0])
                results[line_index][time][3].append(f"")
    results = {key: dict(sorted(value.items(), key=lambda item: item[0])) for key, value in results.items()}
    return results, total


def get_auto_detection_results_from_df_lines(auto_df, min_time=0, max_time=0, lines=None):
    if lines is None:
        raise ValueError("Lines parameter is required for line-based detection results.")
    # lines format: {line_key: list_of_points}
    if max_time == 0:
        max_time = auto_df['time'].max()
    print("Max time set to", max_time)
    results = defaultdict(dict)
    auto_df = auto_df[(auto_df['time'] >= min_time) & (auto_df['time'] <= max_time)]
    THRESHOLD = 0.5
    auto_df = auto_df[auto_df["confidence"] >= THRESHOLD]
    auto_df = auto_df.sort_values(["time", "track_id"])
    grouped = auto_df.groupby('track_id')['time'].agg(['min', 'max'])
    grouped['duration'] = grouped["max"] - grouped["min"]
    grouped = grouped.reset_index()


    grouped = grouped[grouped["duration"] > 2]
    track_ids_of_interest = grouped["track_id"].unique()

    filtered_df = auto_df[auto_df['track_id'].isin(track_ids_of_interest)]
    groups = filtered_df.groupby("track_id")
    total = 0
    for _, group in groups:
        # print("_", _)
        # Sort by time to ensure correct order
        group = group.sort_values("time")
        # Make sure the max time - min time > 2 seconds
        max_time = group["time"].max()
        min_time = group["time"].min()
        if (max_time - min_time) <= 2:
            continue
        xc = (group["x1"] + group["x2"]) / 2
        yc = (group["y1"] + group["y2"]) / 2
        time = group["time"].max()
        final_line_key = None
        final_line_key_mean_degree_angle = float("inf")
        for line_key, line_points in lines.items():


            line_x = line_points["x"]
            line_y = line_points["y"]
            mean_degree_angle = curve_parallelity(xc.values, yc.values, np.array(line_x), np.array(line_y))
            if abs(mean_degree_angle) > 45:
                continue
            
            if mean_degree_angle < final_line_key_mean_degree_angle:
                final_line_key_mean_degree_angle = mean_degree_angle
                final_line_key = line_key
        
        
        if final_line_key is not None:
            
            if time not in results[final_line_key]:
                results[final_line_key][time] = [0, [], [], []]  
            results[final_line_key][time][0] += 1
            total += 1
            results[final_line_key][time][1].append(int(group["track_id"].iloc[0]))
            results[final_line_key][time][2].append(group["cls_id"].iloc[0])
            results[final_line_key][time][3].append(f"{final_line_key_mean_degree_angle:.2f}")
    results = {key: dict(sorted(value.items(), key=lambda item: item[0])) for key, value in results.items()}
    return results, total


def compute_tangents(x, y):
    x = np.asarray(x); y = np.asarray(y)
    dx = np.diff(x)
    dy = np.diff(y)
    tangents = np.stack([dx, dy], axis=1)
    norms = np.linalg.norm(tangents, axis=1, keepdims=True)
    return tangents / (norms + 1e-9)

def curve_parallelity(x1, y1, x2, y2):
    if len(x1) != len(y1) or len(x2) != len(y2) or len(x1) < 2:
        raise ValueError("Input curves must have the same length and contain at least two points.")
    resample_size = max(len(x1), len(x2))
    x1, y1 = resample_points(x1, y1, resample_size)
    x2, y2 = resample_points(x2, y2, resample_size)
    T1 = compute_tangents(x1, y1)
    T2 = compute_tangents(x2, y2)

    cos_sim = np.sum(T1 * T2, axis=1)
    
    angles = np.arccos(cos_sim)

    mean_angle = np.mean(angles)
    # mean angle in degrees
    mean_angle = np.degrees(mean_angle)
    return mean_angle


def resample_points(x, y, num_points):
    x = np.asarray(x, dtype=np.float32)
    y = np.asarray(y, dtype=np.float32)

    n = len(x)
    m = len(y)
    if n != m or n < 2:
        raise ValueError("x and y must have the same length and contain at least two points.")
    
    if n == num_points:
        return x, y
    
    pts = np.stack([x, y], axis=1)
    diffs = np.diff(pts, axis=0)
    seg_lengths = np.sqrt((diffs**2)).sum(axis=1)
    cumdist = np.insert(np.cumsum(seg_lengths), 0, 0)

    new_dist = np.linspace(0, cumdist[-1], num_points)
    new_x = np.interp(new_dist, cumdist, x)
    new_y = np.interp(new_dist, cumdist, y)
    return new_x, new_y

def _segments_intersect(p1, p2, p3, p4):
    """Return True if segment p1->p2 intersects segment p3->p4."""
    def _cross2d(a, b):
        return a[0] * b[1] - a[1] * b[0]
    d1 = (p2[0] - p1[0], p2[1] - p1[1])
    d2 = (p4[0] - p3[0], p4[1] - p3[1])
    denom = _cross2d(d1, d2)
    if abs(denom) < 1e-10:
        return False
    diff = (p3[0] - p1[0], p3[1] - p1[1])
    t = _cross2d(diff, d2) / denom
    u = _cross2d(diff, d1) / denom
    return 0.0 <= t <= 1.0 and 0.0 <= u <= 1.0


def get_auto_detection_results_from_df_crossings(auto_df, min_time=0, max_time=0, crossing_lines=None):
    """
    v5 aggregation: for each track, walk its centroid path frame-by-frame and record
    the first time it crosses each drawn crossing_line segment.
    crossing_lines: {line_key: {"x": [x1, x2], "y": [y1, y2]}}  (normalized 0-1 coords)
    """
    if crossing_lines is None:
        raise ValueError("crossing_lines required for v5")
    if max_time == 0:
        max_time = auto_df["time"].max()
    results = defaultdict(dict)
    auto_df = auto_df[(auto_df["time"] >= min_time) & (auto_df["time"] <= max_time)]
    auto_df = auto_df[auto_df["confidence"] >= 0.3]
    auto_df = auto_df.sort_values(["track_id", "time"])

    dur = auto_df.groupby("track_id")["time"].agg(["min", "max"])
    dur["duration"] = dur["max"] - dur["min"]
    valid_ids = dur[dur["duration"] > 2].index
    auto_df = auto_df[auto_df["track_id"].isin(valid_ids)]

    total = 0
    crossed = set()  # (track_id, line_key) — count each pair only once

    for track_id, group in auto_df.groupby("track_id"):
        group = group.sort_values("time")
        cx = ((group["x1"] + group["x2"]) / 2).values
        cy = ((group["y1"] + group["y2"]) / 2).values
        times = group["time"].values
        cls_ids = group["cls_id"].values

        for i in range(1, len(cx)):
            p1 = (cx[i - 1], cy[i - 1])
            p2 = (cx[i], cy[i])
            for line_key, line_data in crossing_lines.items():
                if (track_id, line_key) in crossed:
                    continue
                xs = line_data.get("x", [])
                ys = line_data.get("y", [])
                if len(xs) < 2 or len(ys) < 2:
                    continue
                p3 = (xs[0], ys[0])
                p4 = (xs[-1], ys[-1])
                if _segments_intersect(p1, p2, p3, p4):
                    crossed.add((track_id, line_key))
                    t = float(times[i])
                    if t not in results[line_key]:
                        results[line_key][t] = [0, [], [], []]
                    results[line_key][t][0] += 1
                    results[line_key][t][1].append(int(track_id))
                    results[line_key][t][2].append(int(cls_ids[i]))
                    results[line_key][t][3].append("")
                    total += 1

    results = {k: dict(sorted(v.items())) for k, v in results.items()}
    return results, total


def get_counter_auto_detection_results(record_id, version, divide_time, min_time=0, max_time=0):
    """
    API endpoint to retrieve auto_detection counting results for a specific counter.
    Expects a GET request with 'counter_id' as a query parameter.
    """
    if not record_id:
        print("record_id is required")
        return False, 0
    record = Record.objects.filter(id=record_id).first()
    if not record:
        print("record not found")
        return False, 0
    auto_detection = AutoDetection.objects.filter(record=record, version=version, divide_time=divide_time).first()
    if not auto_detection:
        print("auto_detection not found")
        return False, 0
    counts_file = auto_detection.file_name
    if not os.path.exists(counts_file):
        print("counts_file not found")
        return False, 0
    try:
        detection_lines = DetectionLines.objects.filter(record=record).first()
        if not detection_lines:
            print("detection_lines not found")
            return False, 0
        lines = detection_lines.lines
        lines_map_length = {
            zone_name: len(list(filter(lambda x: x["tool"] == "zone", list_of_points))) for zone_name, list_of_points in lines.items()
        }
        # lines format: {line_name: [{"tool": "zone"/"direction", "points": [...]}, ...]}
        # points format: [x1, y1, x2, y2, ...]
        # we want: {line_name: {"x": [...], "y": [...]} }
        lines_direction = {
            line_name: {
                "x": [point for tool in list_of_points for idx, point in enumerate(tool["points"]) if idx % 2 == 0 and tool["tool"] == "direction"],
                "y": [point for tool in list_of_points for idx, point in enumerate(tool["points"]) if idx % 2 == 1 and tool["tool"] == "direction"],
            } for line_name, list_of_points in lines.items()
        }
        df = pd.read_csv(counts_file)
        if version == "v3" or version == "v2":
            return get_auto_detection_results_from_df_zones(df, min_time, max_time, lines_map_length)
        elif version == "v4":
            return get_auto_detection_results_from_df_lines(df, min_time, max_time, lines_direction)
        elif version == "v5":
            lines_crossing = {
                line_name: {
                    "x": [pt for tool in list_of_points for idx, pt in enumerate(tool["points"]) if idx % 2 == 0 and tool["tool"] == "crossing_line"],
                    "y": [pt for tool in list_of_points for idx, pt in enumerate(tool["points"]) if idx % 2 == 1 and tool["tool"] == "crossing_line"],
                }
                for line_name, list_of_points in lines.items()
            }
            return get_auto_detection_results_from_df_crossings(df, min_time, max_time, lines_crossing)
        return False, 0
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return False, 0


def get_counter_manual_results(record_id,min_time=0, max_time=0):
    """
    API endpoint to retrieve manual counting results for a specific counter.
    """
    record = Record.objects.filter(id=record_id).first()
    if not record:
        return False, 0
    
    record_logs = RecordLog.objects.filter(record=record).order_by('time')
        
    if not record_logs.exists():
        return False, 0

    # Making a pandas DataFrame from the record logs
    df = pd.DataFrame(list(record_logs.values('time', 'turn_movement')), columns=['time', 'turn_movement'])
    
    if max_time == 0:
        max_time = float("inf")
    df = df[(df["time"] >= min_time) & (df["time"] <= max_time)]

    df = df.sort_values(["time"])
    results = defaultdict(dict)
    groups = df.groupby('turn_movement')
    total = 0
    for turn_movement, group in groups:
        for _, row in group.iterrows():
            time = row['time']
            if time not in results[turn_movement]:
                results[turn_movement][time] = 0
            results[turn_movement][time] += 1
            total += 1

    results = {key: dict(sorted(value.items(), key=lambda item: item[0])) for key, value in results.items()}
    return results, total

def get_iss_detections_json(record_id, min_time=0, max_time=0):
    """
    using the arguments to form a url like: http://192.168.42.169/api/v1/cameras/1/detections?start-time=2025-05-10T12:00:00&end-time=2025-05-11T13:00:00
    min_time: the minimum `seconds` that should be added to the start time of the recording
    max_time: the maximum `seconds` that should be added to the end time of the record
    """
    record = Record.objects.filter(id=record_id).first()
    if not record:
        return False
    ip = get_ip_from_rtsp(record.camera_url)
    record_start_time = record.start_time
    record_start_time = record_start_time + timedelta(seconds=min_time)
    # TODO check this part right here:
    if max_time == 0:
        max_time = record.duration * 60
    record_end_time = record.start_time + timedelta(seconds=max_time)
    url = f"http://{ip}/api/v1/cameras/{record.camera_id}/detections"
    params = {
        "start-time": record_start_time.strftime("%Y-%m-%dT%H:%M:%S"),
        "end-time": record_end_time.strftime("%Y-%m-%dT%H:%M:%S")
    }
    try:
        response = requests.get(url, params=params)
        response.raise_for_status() 
        return response.json()["detections"] 
    except requests.RequestException as e:
        print(f"Error fetching detections: {e}")
        return None

def get_iss_detections_pandas(record_id, min_time=0, max_time=0):
    """
    Get the detections from this class using get_detections_json but format it into a pandas dataframe
    """
    data_json = get_iss_detections_json(record_id, min_time, max_time)

    if data_json is None:
        return pd.DataFrame(), 0
    pandas_df = pd.DataFrame(data_json)
    
    if pandas_df.empty:
        return pandas_df, 0
    pandas_df["direction"] = pandas_df["direction"].apply(
        lambda x: "through" if x == "Through" else "left" if x == "LeftTurn" else "right" if x == "RightTurn" else x
    )
    pandas_df = pandas_df[~pandas_df["zoneName"].str.contains("ADV")]
    total = pandas_df.shape[0]
    return pandas_df, total

def get_results_comparison_df(record_id, version, divide_time, min_time=0, max_time=0):
    manual_counts, manaul_total = get_counter_manual_results(record_id, min_time=min_time, max_time=max_time)
    auto_counts, auto_total = get_counter_auto_detection_results(record_id, version, divide_time, min_time=min_time, max_time=max_time)
    iss_api_df, iss_total = get_iss_detections_pandas(record_id, min_time, max_time)
    return (manual_counts, manaul_total), (auto_counts, auto_total), (iss_api_df, iss_total)



def get_counting_raw(df):
    
    if df.empty:
        return pd.DataFrame()
    columns = ["time", "count"]
    for column in columns:
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found in DataFrame.")
    
    # time column should be datatype (datetime64[ns])
    if not pd.api.types.is_datetime64_any_dtype(df["time"]):
        df["time"] = pd.to_datetime(df["time"]).dt.tz_localize(None)
    
    df["time"] = pd.to_datetime(df["time"]).dt.tz_localize(None)
    
    # Divide the whole day into 15-minute intervals. Beginning from 00:00 to 23.45. But we need to
    # keep this in the US format: PM and AM
    df["hour"] = df["time"].dt.hour
    df["minute"] = df["time"].dt.minute
    start = df["time"].min().floor("15min")
    end = df["time"].max().ceil("15min")
    df["interval"] = df["time"].dt.floor("15min")
    counts = (
        df.groupby("interval")
        .size()
        .reindex(pd.date_range(start, end, freq="15min"), fill_value=0)
    )
    
    result = counts.reset_index()
    result.columns = ["interval_start", "count"]
    result["interval_start"] = result["interval_start"].dt.tz_localize("UTC").dt.tz_convert("America/Chicago")
    result["label"] = result["interval_start"].dt.strftime("%I:%M %p")
    return result


def get_manual_counting_excel(record_id):

    record = Record.objects.filter(id=record_id).first()
    if not record:
        raise ValueError("Record not found.")
    start_time = record.start_time
    manual_results, manual_total = get_counter_manual_results(record_id)
    final_results = pd.DataFrame({"interval_start":[]})
    for line_key, line_data in manual_results.items():
        results_before_df = {"time":[], "count":[]}
        for time, count in line_data.items():
            results_before_df["time"].append(start_time + timedelta(seconds=time))
           # print(start_time + timedelta(seconds=time)) 
            results_before_df["count"].append(count)
        results_before_df["time"] = (
            pd.Series(results_before_df["time"])
            .dt.tz_localize(None)
        )
        results_df_for_line_key = pd.DataFrame(results_before_df)
        raw_results_df_for_line_key = get_counting_raw(results_df_for_line_key)
        raw_results_df_for_line_key = raw_results_df_for_line_key.rename(
            columns={"count": line_key + " Count"}
        )
        raw_results_df_for_line_key["interval_start"] = pd.to_datetime(raw_results_df_for_line_key["interval_start"])
        final_results = pd.merge(
            final_results,
            raw_results_df_for_line_key[["interval_start", line_key + " Count"]],
            on="interval_start",
            how="outer"
        )
        
    
    results = final_results.sort_values("interval_start")
    numeric_cols = results.select_dtypes(include="number").columns
    results[numeric_cols] = results[numeric_cols].fillna(0)
    return results

def get_auto_detection_counting_excel(record_id, version, divide_time):
    auto_results, total = get_counter_auto_detection_results(record_id, version, divide_time)
    if not auto_results:
        raise ValueError("No auto detection results found.")
    record = Record.objects.filter(id=record_id).first()
    if not record:
        raise ValueError("Record not found.")
    start_time = record.start_time
    results = pd.DataFrame({"interval_start":[]})
    
    for line_key, line_data in auto_results.items():
        results_before_df = {"time":[], "count":[]}
        for time, data in line_data.items():
            results_before_df["time"].append(start_time + timedelta(seconds=time))
            results_before_df["count"].append(data[0])
        results_df_for_line_key = pd.DataFrame(results_before_df)
        raw_results_df_for_line_key = get_counting_raw(results_df_for_line_key)
        raw_results_df_for_line_key = raw_results_df_for_line_key.rename(
            columns={"count": line_key + " Count"}
        )
        results = pd.merge(
            results,
            raw_results_df_for_line_key[["interval_start", line_key + " Count"]],
            on="interval_start",
            how="outer"
        )
    
    results = results.sort_values("interval_start").fillna(0)
    return results

def get_iss_detections_counting_excel(record_id, min_time=0, max_time=0):
    iss_api_df, _ = get_iss_detections_pandas(record_id, min_time, max_time)
    if iss_api_df.empty:
        raise ValueError("No ISS API detection results found.")
    record = Record.objects.filter(id=record_id).first()
    if not record:
        raise ValueError("Record not found.")
    start_time = record.start_time
    results = pd.DataFrame({"interval_start":[]})
    movement_groups = iss_api_df.groupby("direction")
    for movement, group in movement_groups:
        results_before_df = {"time":[], "count":[]}
        time_counts = group['time'].value_counts().sort_index()
        for time, count in time_counts.items():
            results_before_df["time"].append(start_time)
            results_before_df["count"].append(count)
        results_df_for_movement = pd.DataFrame(results_before_df)
        raw_results_df_for_movement = get_counting_raw(results_df_for_movement)
        raw_results_df_for_movement = raw_results_df_for_movement.rename(
            columns={"count": movement + " Count"}
        )
        results = pd.merge(
            results,
            raw_results_df_for_movement[["interval_start", movement + " Count"]],
            on="interval_start",
            how="outer"
        )
        
    results = results.sort_values("interval_start").fillna(0)
    return results


def get_manual_count_excel_with_direction(record_id):
    """
    Write manual counting results to an Excel file with a two-row header:
    - Row 1: empty first cell, then a merged cell with the direction spanning all data columns
    - Row 2: actual column names (time, <Direction> Count, ...)
    This avoids pandas' MultiIndex->Excel limitation by writing headers with openpyxl.
    """
    record = Record.objects.filter(id=record_id).first()
    if not record:
        raise ValueError("Record not found.")
    direction = record.direction or "N/A"

    excel_df = get_manual_counting_excel(record_id)
    if excel_df is None or excel_df.empty:
        return excel_df

    # Ensure 'time' column name (some helpers call it 'interval_start')
    # prefer 'time' if present, otherwise rename 'interval_start' -> 'time' for output
    if "interval_start" in excel_df.columns and "time" not in excel_df.columns:
        excel_df = excel_df.rename(columns={"interval_start": "time"})

    # Prepare header groups: we'll put the 'direction' header above all columns except the first ('time')
    cols = excel_df.columns.tolist()
    if len(cols) < 1:
        raise ValueError("DataFrame has no columns to write.")

    # Build workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1: first cell empty, merge remaining columns under direction
    first_data_col = 2  # column 1 will be 'time'
    last_data_col = first_data_col + (len(cols) - 1) - 1 + 1  # compute inclusive last column index
    # write empty top-left cell
    ws.cell(row=1, column=1, value="")
    # merge and write direction if there are data columns
    if len(cols) > 1:
        start_col_letter = get_column_letter(first_data_col)
        end_col_letter = get_column_letter(first_data_col + len(cols) - 2)
        merge_range = f"{start_col_letter}1:{end_col_letter}1"
        ws.merge_cells(merge_range)
        ws[f"{start_col_letter}1"] = direction
        ws[f"{start_col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        # Only time column -> write direction above time (even though user wanted empty top-left,
        # in case of single column we keep top-left empty)
        pass

    # Row 2: write actual column names
    for j, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=j, value=str(col_name))
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3+: write data rows
    # Ensure pandas datetimes become python datetimes
    excel_df_copy = excel_df.copy()
    for c in excel_df_copy.select_dtypes(include=["datetime", "datetimetz"]).columns:
        excel_df_copy[c] = pd.to_datetime(excel_df_copy[c]).dt.tz_localize(None)

    for i, row in enumerate(excel_df_copy.itertuples(index=False), start=3):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    # Optionally set column widths
    for i, col in enumerate(cols, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(12, min(30, len(str(col)) + 2))
    return wb


def get_auto_iss_count_excel_with_direction(record_id, version, divide_time):
    """
    Get an excel with two sheets:
    -- First sheet is auto excel results from get_auto_detection_counting_excel
    -- Seconds sheet is ISS excel results from get_iss_detections_counting_excel
    This avoids pandas' MultiIndex->Excel limitation by writing headers with openpyxl.
    """
    record = Record.objects.filter(id=record_id).first()
    if not record:
        raise ValueError("Record not found.")
    direction = record.direction or "N/A"
    finished_detecting = record.finished_detecting
    auto_excel_df = get_auto_detection_counting_excel(record_id, version, divide_time)
    iss_excel_df = get_iss_detections_counting_excel(record_id)
    
    if auto_excel_df is None or auto_excel_df.empty:
        raise ValueError("No auto detection counting results found.")
    if iss_excel_df is None or iss_excel_df.empty:
        raise ValueError("No ISS detection counting results found.")

    # Prepare workbook
    wb = openpyxl.Workbook()
    ws_auto = wb.active
    ws_auto.title = "Auto Detection Counts"
    ws_iss = wb.create_sheet(title="ISS Detection Counts")

    # Helper to write a DataFrame to a worksheet with two-row header
    def write_df_with_direction_header(ws, df, direction):
        # Ensure 'time' column name (some helpers call it 'interval_start')
        if "interval_start" in df.columns and "time" not in df.columns:
            df = df.rename(columns={"interval_start": "time"})

        cols = df.columns.tolist()
        if len(cols) < 1:
            raise ValueError("DataFrame has no columns to write.")

        # Row 1: first cell empty, merge remaining columns under direction
        first_data_col = 2  # column 1 will be 'time'
        last_data_col = first_data_col + (len(cols) - 1) - 1 + 1  # compute inclusive last column index
        ws.cell(row=1, column=1, value="")
        if len(cols) > 1:
            start_col_letter = get_column_letter(first_data_col)
            end_col_letter = get_column_letter(first_data_col + len(cols) - 2)
            merge_range = f"{start_col_letter}1:{end_col_letter}1"
            ws.merge_cells(merge_range)
            ws[f"{start_col_letter}1"] = direction
            ws[f"{start_col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")

        # Row 2: write actual column names
        for j, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=j, value=str(col_name))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Row 3+: write data rows
        df_copy = df.copy()
        for c in df_copy.select_dtypes(include=["datetime", "datetimetz"]).columns:
            df_copy[c] = pd.to_datetime(df_copy[c]).dt.tz_localize(None)
        for i, row in enumerate(df_copy.itertuples(index=False), start=3):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
    
    write_df_with_direction_header(ws_auto, auto_excel_df, direction)
    write_df_with_direction_header(ws_iss, iss_excel_df, direction)
    # Get comparison with Actual results
    
    if finished_detecting:
        manual_counts_excel = get_manual_counting_excel(record_id)
        manual_counts_excel = manual_counts_excel.rename(columns={"interval_start": "time", "turn_movement": "portal"})
        # Comparison with Auto
        transposed_manual_counts = transpose_df(manual_counts_excel, direction)
        transposed_auto_counts = transpose_df(auto_excel_df.rename(columns={"interval_start": "time"}), direction)
        transposed_iss_counts = transpose_df(iss_excel_df.rename(columns={"interval_start": "time"}), direction)
        auto_comparison = compare_two_df(transposed_manual_counts, transposed_auto_counts)
        # Comparison with ISS
        iss_comparison = compare_two_df(transposed_manual_counts, transposed_iss_counts)

        # Write comparison sheets
        
        copy_sheet_to_workbook(auto_comparison, wb, "Manual vs Auto Comparison")
        copy_sheet_to_workbook(iss_comparison, wb, "Manual vs ISS Comparison")
        
            
    return wb
    
def transpose_df(df, direction):
    all_dfs_in_one = pd.DataFrame({"time": [], "direction":[], "count":[], "portal": []})
    for col in df.columns:
        if col == "time" or col == "interval_start":
            continue
        temp_df = pd.DataFrame()
        temp_df["time"] = df["time"] if "time" in df.columns else df["interval_start"]
        temp_df["direction"] = direction
        temp_df["portal"] = col
        temp_df["count"] = df[col]
        if not temp_df.empty and not temp_df.isna().all().all():
            all_dfs_in_one = pd.concat([all_dfs_in_one, temp_df], ignore_index=True)
    return all_dfs_in_one

def get_multiple_manual_counting_excel(records_ids):
    """
    Get the same results as get_manual_count_excel_with_direction for each of the records
    and merge them based on the time. They should be like the following:
    | EMPTY | Direction 1 (with a1 columns)| Direction 2 (with a2 columns) | ... | Direction n (an) columns |
    | EMPTY | Count movement 1 | Count movement 2 | ... | Count movement a1  ...| | Count movement 1 | ... | Count movement an |
    | Time | Count | Count | ... | Count |
    ...
    """
    combined_results = pd.DataFrame({"time":[]})
    directions = []
    record_column_info = []  # will store tuples (direction, num_cols, col_names)
    all_dfs = {}
    all_dfs_in_one = pd.DataFrame()

    for record_id in records_ids:
        record = Record.objects.filter(id=record_id).first()
        if not record:
            continue
        direction = record.direction or f"N/A direction of record_id {record_id}"
        
        excel_df = get_manual_counting_excel(record_id)
        excel_transposed_df = transpose_df(excel_df, direction)
        all_dfs_in_one = pd.concat([all_dfs_in_one, excel_transposed_df], ignore_index=True)
        all_dfs[record_id] = excel_df
        if excel_df is None or excel_df.empty:
            continue
        if "interval_start" in excel_df.columns and "time" not in excel_df.columns:
            excel_df = excel_df.rename(columns={"interval_start": "time"})
        # count data columns for this record (exclude 'time')
        data_cols = [c for c in excel_df.columns if c != "time"]
        record_column_info.append((direction, len(data_cols), data_cols, record_id))
        directions.append(direction)
        # Rename columns to include direction to avoid collisions
        renamed_columns = {
            col: f"{direction} {col}" if col != "time" else col for col in excel_df.columns
        }
        excel_df = excel_df.rename(columns=renamed_columns)
        combined_results = pd.merge(
            combined_results,
            excel_df,
            on="time",
            how="outer"
        )
    if combined_results.empty:
        raise ValueError("No valid manual counting results found for the provided record IDs.")

    # ensure time is datetime and sort
    if "time" in combined_results.columns:
        combined_results["time"] = pd.to_datetime(combined_results["time"])
    combined_results = combined_results.sort_values("time").reset_index(drop=True)

    # Now write to Excel with two-row header and styling
    wb = openpyxl.Workbook()
    ws = wb.active

    # quick style objects
    from openpyxl.styles import PatternFill, Border, Side, Font
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    time_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
    fills = [
        PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # light blue
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # light gray
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # light green
        PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # light peach
    ]

    # Row 1: first cell empty
    ws.cell(row=1, column=1, value="")
    # Determine columns per direction group (order in combined_results.columns)
    cols = combined_results.columns.tolist()
    # Build mapping of direction -> (start_col_index, end_col_index, num_cols, record_id)
    mapping = []
    col_index = 1
    # find indices for 'time' then subsequent columns
    # ensure time is first; if not, move it to first in ordering
    if cols[0] != "time" and "time" in cols:
        cols.remove("time")
        cols.insert(0, "time")
    # calculate contiguous groups for each record using record_column_info order
    current_col = 2  # excel column index for first data column
    for direction, num_cols, data_cols, record_id in record_column_info:
        if num_cols == 0:
            continue
        # find the actual column names in combined_results corresponding to this direction
        full_col_names = [f"{direction} {c}" for c in data_cols]
        # find contiguous block start by searching for first occurrence among cols
        found_indices = [cols.index(name) + 1 for name in full_col_names if name in cols]
        if not found_indices:
            continue
        start_idx = min(found_indices)
        end_idx = max(found_indices)
        # store mapping as 1-based excel col indices
        mapping.append((direction, start_idx, end_idx, num_cols, record_id))
    # If mapping empty, try fallback: group by prefix direction in column order
    if not mapping:
        # scan cols and group by prefix before first space
        idx = 2
        seen_dirs = set()
        for c in cols[1:]:
            parts = str(c).split(" ", 1)
            dir_name = parts[0]
            if dir_name not in seen_dirs:
                # count how many consecutive columns start with this dir_name
                group_cols = [cc for cc in cols[1:] if str(cc).startswith(dir_name + " ")]
                if not group_cols:
                    continue
                start_idx = cols.index(group_cols[0]) + 1
                end_idx = cols.index(group_cols[-1]) + 1
                mapping.append((dir_name, start_idx, end_idx, len(group_cols), None))
                seen_dirs.add(dir_name)

    # Row 1: merge and write direction headers with alternating fills and borders per record block
    for idx, (direction, start_idx, end_idx, num_cols, record_id) in enumerate(mapping):
        start_letter = get_column_letter(start_idx)
        end_letter = get_column_letter(end_idx)
        merge_range = f"{start_letter}1:{end_letter}1"
        ws.merge_cells(merge_range)
        cell = ws.cell(row=1, column=start_idx, value=direction)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        fill = fills[idx % len(fills)]
        # apply fill to merged header range
        for col in range(start_idx, end_idx + 1):
            hcell = ws.cell(row=1, column=col)
            hcell.fill = fill
            hcell.border = border
            hcell.font = Font(bold=True)
        # set column widths: total width 90 spread over the group's columns
        per_col_width = 90.0 / max(1, num_cols)
        for col in range(start_idx, end_idx + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = per_col_width

    # time column styling and width
    time_col_idx = cols.index("time") + 1 if "time" in cols else 1
    time_letter = get_column_letter(time_col_idx)
    ws.cell(row=1, column=time_col_idx).fill = time_fill  # top-left may be empty but color the cell
    ws.cell(row=2, column=time_col_idx).fill = time_fill
    ws.column_dimensions[time_letter].width = 20

    # Row 2: actual column names
    for j, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=j, value=str(col_name))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # apply time fill
        if j == time_col_idx:
            cell.fill = time_fill
        else:
            # find which record block this column belongs to and apply corresponding fill
            for idx, (direction, start_idx, end_idx, num_cols, record_id) in enumerate(mapping):
                if start_idx <= j <= end_idx:
                    cell.fill = fills[idx % len(fills)]
                    break
        cell.border = border

    # Row 3+: data rows
    combined_results_copy = combined_results.copy()
    for c in combined_results_copy.select_dtypes(include=["datetime", "datetimetz"]).columns:
        combined_results_copy[c] = pd.to_datetime(combined_results_copy[c]).dt.tz_localize(None)
    n_rows = combined_results_copy.shape[0]
    for i, row in enumerate(combined_results_copy.itertuples(index=False), start=3):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            # apply border
            cell.border = border
            # apply time column fill
            if j == time_col_idx:
                cell.fill = time_fill
            else:
                # fill by record block
                for idx, (direction, start_idx, end_idx, num_cols, record_id) in enumerate(mapping):
                    if start_idx <= j <= end_idx:
                        cell.fill = fills[idx % len(fills)]
                        break

    # Also add outer border around each record block (including header rows + data rows)
    last_data_row = 2 + n_rows
    for idx, (_, start_idx, end_idx, _, _) in enumerate(mapping):
        for r in range(1, last_data_row + 1):
            for c in range(start_idx, end_idx + 1):
                ws.cell(row=r, column=c).border = border
    return wb, all_dfs_in_one


def get_multiple_iss_counting_excel(records_ids):
    """
    Get the same results as get_manual_count_excel_with_direction for each of the records
    and merge them based on the time. They should be like the following:
    | EMPTY | Direction 1 (with a1 columns)| Direction 2 (with a2 columns) | ... | Direction n (an) columns |
    | EMPTY | Count movement 1 | Count movement 2 | ... | Count movement a1  ...| | Count movement 1 | ... | Count movement an |
    | Time | Count | Count | ... | Count |
    ...
    """
    combined_results = pd.DataFrame({"time":[]})
    directions = []
    record_column_info = []  # will store tuples (direction, num_cols, col_names)
    all_dfs = {}
    all_dfs_in_one = pd.DataFrame()
    for record_id in records_ids:
        record = Record.objects.filter(id=record_id).first()
        if not record:
            continue
        direction = record.direction or f"N/A direction of record_id {record_id}"
        excel_df = get_iss_detections_counting_excel(record_id)
        excel_transposed_df = transpose_df(excel_df, direction)
        all_dfs_in_one = pd.concat([all_dfs_in_one, excel_transposed_df], ignore_index=True)
        all_dfs[record_id] = excel_df
        if excel_df is None or excel_df.empty:
            continue
        if "interval_start" in excel_df.columns and "time" not in excel_df.columns:
            excel_df = excel_df.rename(columns={"interval_start": "time"})
        # count data columns for this record (exclude 'time')
        data_cols = [c for c in excel_df.columns if c != "time"]
        record_column_info.append((direction, len(data_cols), data_cols, record_id))
        directions.append(direction)
        # Rename columns to include direction to avoid collisions
        renamed_columns = {
            col: f"{direction} {col}" if col != "time" else col for col in excel_df.columns
        }
        excel_df = excel_df.rename(columns=renamed_columns)
        combined_results = pd.merge(
            combined_results,
            excel_df,
            on="time",
            how="outer"
        )
    if combined_results.empty:
        raise ValueError("No valid manual counting results found for the provided record IDs.")

    # ensure time is datetime and sort
    if "time" in combined_results.columns:
        combined_results["time"] = pd.to_datetime(combined_results["time"])
    combined_results = combined_results.sort_values("time").reset_index(drop=True)

    # Now write to Excel with two-row header and styling
    wb = openpyxl.Workbook()
    ws = wb.active

    # quick style objects
    from openpyxl.styles import PatternFill, Border, Side, Font
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    time_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
    fills = [
        PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # light blue
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # light gray
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # light green
        PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # light peach
    ]

    # Row 1: first cell empty
    ws.cell(row=1, column=1, value="")
    # Determine columns per direction group (order in combined_results.columns)
    cols = combined_results.columns.tolist()
    # Build mapping of direction -> (start_col_index, end_col_index, num_cols, record_id)
    mapping = []
    col_index = 1
    # find indices for 'time' then subsequent columns
    # ensure time is first; if not, move it to first in ordering
    if cols[0] != "time" and "time" in cols:
        cols.remove("time")
        cols.insert(0, "time")
    # calculate contiguous groups for each record using record_column_info order
    current_col = 2  # excel column index for first data column
    for direction, num_cols, data_cols, record_id in record_column_info:
        if num_cols == 0:
            continue
        # find the actual column names in combined_results corresponding to this direction
        full_col_names = [f"{direction} {c}" for c in data_cols]
        # find contiguous block start by searching for first occurrence among cols
        found_indices = [cols.index(name) + 1 for name in full_col_names if name in cols]
        if not found_indices:
            continue
        start_idx = min(found_indices)
        end_idx = max(found_indices)
        # store mapping as 1-based excel col indices
        mapping.append((direction, start_idx, end_idx, num_cols, record_id))
    # If mapping empty, try fallback: group by prefix direction in column order
    if not mapping:
        # scan cols and group by prefix before first space
        idx = 2
        seen_dirs = set()
        for c in cols[1:]:
            parts = str(c).split(" ", 1)
            dir_name = parts[0]
            if dir_name not in seen_dirs:
                # count how many consecutive columns start with this dir_name
                group_cols = [cc for cc in cols[1:] if str(cc).startswith(dir_name + " ")]
                if not group_cols:
                    continue
                start_idx = cols.index(group_cols[0]) + 1
                end_idx = cols.index(group_cols[-1]) + 1
                mapping.append((dir_name, start_idx, end_idx, len(group_cols), None))
                seen_dirs.add(dir_name)

    # Row 1: merge and write direction headers with alternating fills and borders per record block
    for idx, (direction, start_idx, end_idx, num_cols, record_id) in enumerate(mapping):
        start_letter = get_column_letter(start_idx)
        end_letter = get_column_letter(end_idx)
        merge_range = f"{start_letter}1:{end_letter}1"
        ws.merge_cells(merge_range)
        cell = ws.cell(row=1, column=start_idx, value=direction)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        fill = fills[idx % len(fills)]
        # apply fill to merged header range
        for col in range(start_idx, end_idx + 1):
            hcell = ws.cell(row=1, column=col)
            hcell.fill = fill
            hcell.border = border
            hcell.font = Font(bold=True)
        # set column widths: total width 90 spread over the group's columns
        per_col_width = 90.0 / max(1, num_cols)
        for col in range(start_idx, end_idx + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = per_col_width

    # time column styling and width
    time_col_idx = cols.index("time") + 1 if "time" in cols else 1
    time_letter = get_column_letter(time_col_idx)
    ws.cell(row=1, column=time_col_idx).fill = time_fill  # top-left may be empty but color the cell
    ws.cell(row=2, column=time_col_idx).fill = time_fill
    ws.column_dimensions[time_letter].width = 20

    # Row 2: actual column names
    for j, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=j, value=str(col_name))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # apply time fill
        if j == time_col_idx:
            cell.fill = time_fill
        else:
            # find which record block this column belongs to and apply corresponding fill
            for idx, (direction, start_idx, end_idx, num_cols, record_id) in enumerate(mapping):
                if start_idx <= j <= end_idx:
                    cell.fill = fills[idx % len(fills)]
                    break
        cell.border = border

    # Row 3+: data rows
    combined_results_copy = combined_results.copy()
    for c in combined_results_copy.select_dtypes(include=["datetime", "datetimetz"]).columns:
        combined_results_copy[c] = pd.to_datetime(combined_results_copy[c]).dt.tz_localize(None)
    n_rows = combined_results_copy.shape[0]
    for i, row in enumerate(combined_results_copy.itertuples(index=False), start=3):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            # apply border
            cell.border = border
            # apply time column fill
            if j == time_col_idx:
                cell.fill = time_fill
            else:
                # fill by record block
                for idx, (direction, start_idx, end_idx, num_cols, record_id) in enumerate(mapping):
                    if start_idx <= j <= end_idx:
                        cell.fill = fills[idx % len(fills)]
                        break

    # Also add outer border around each record block (including header rows + data rows)
    last_data_row = 2 + n_rows
    for idx, (_, start_idx, end_idx, _, _) in enumerate(mapping):
        for r in range(1, last_data_row + 1):
            for c in range(start_idx, end_idx + 1):
                ws.cell(row=r, column=c).border = border
    return wb, all_dfs_in_one



def get_multiple_auto_counting_excel(records_ids):
    """
    Aggregate auto-detection counting results by (version, divide_time).
    Produce one sheet per unique (version, divide_time) containing merged
    results for records that share that pair. Each sheet groups columns
    by record.direction (same styling as manual export).
    """
    # gather dfs per (version, divide_time)
    groups = {}  # (version, divide_time) -> list of tuples (record_id, direction, df)
    all_dfs = {}
    all_dfs_in_one = pd.DataFrame()
    for record_id in records_ids:
        auto_detection = AutoDetection.objects.filter(record_id=record_id).order_by("-id").first()
    
        version = auto_detection.version
        divide_time = auto_detection.divide_time
        try:
            excel_df = get_auto_detection_counting_excel(record_id, version, divide_time)
            all_dfs_in_one = pd.concat([all_dfs_in_one, transpose_df(excel_df, f"Record {record_id}")], ignore_index=True)
            all_dfs[str(record_id) + "_" + str(version) + "_" + str(divide_time)] = excel_df
            

        except Exception:
            excel_df = None
        if excel_df is None or excel_df.empty:
            continue
        # normalize time column name
        if "interval_start" in excel_df.columns and "time" not in excel_df.columns:
            excel_df = excel_df.rename(columns={"interval_start": "time"})
        record = Record.objects.filter(id=record_id).first()
        direction = (record.direction or "N/A") if record else "N/A"
        key = (version, divide_time)
        groups.setdefault(key, []).append((record_id, direction, excel_df.copy()))

    if not groups:
        raise ValueError("No valid auto detection results found for the provided record IDs.")

    wb = openpyxl.Workbook()
    # keep default sheet to remove later if unused
    default_sheet = wb.active
    created_any = False

    # styling helpers reused from manual exporter
    from openpyxl.styles import PatternFill, Border, Side, Font
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    time_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fills = [
        PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    ]

    def write_group_sheet(title, records_list):
        """
        records_list: list of (record_id, direction, df)
        returns: created worksheet
        """
        combined = pd.DataFrame({"time": []})
        record_column_info = []
        directions = []
        for record_id, direction, df in records_list:
            # ensure time column present
            if "interval_start" in df.columns and "time" not in df.columns:
                df = df.rename(columns={"interval_start": "time"})
            data_cols = [c for c in df.columns if c != "time"]
            record_column_info.append((direction, len(data_cols), data_cols, record_id))
            directions.append(direction)
            # rename to avoid collisions
            renamed = {col: f"{direction} {col}" if col != "time" else col for col in df.columns}
            df2 = df.rename(columns=renamed)
            combined = pd.merge(combined, df2, on="time", how="outer")
        if combined.empty:
            return None

        # normalize time dtype and sort
        if "time" in combined.columns:
            combined["time"] = pd.to_datetime(combined["time"])
        combined = combined.sort_values("time").reset_index(drop=True)

        ws = wb.create_sheet(title=title[:31])
        # Row1 blank cell
        ws.cell(row=1, column=1, value="")
        cols = combined.columns.tolist()
        # ensure time first
        if cols and cols[0] != "time" and "time" in cols:
            cols.remove("time"); cols.insert(0, "time")

        # build mapping as in manual function
        mapping = []
        for direction, num_cols, data_cols, record_id in record_column_info:
            if num_cols == 0:
                continue
            full_col_names = [f"{direction} {c}" for c in data_cols]
            found_indices = [cols.index(name) + 1 for name in full_col_names if name in cols]
            if not found_indices:
                continue
            mapping.append((direction, min(found_indices), max(found_indices), num_cols, record_id))
        if not mapping:
            # fallback grouping by prefix
            seen = set()
            for c in cols[1:]:
                prefix = str(c).split(" ", 1)[0]
                if prefix in seen:
                    continue
                group_cols = [cc for cc in cols[1:] if str(cc).startswith(prefix + " ")]
                if not group_cols:
                    continue
                start_idx = cols.index(group_cols[0]) + 1
                end_idx = cols.index(group_cols[-1]) + 1
                mapping.append((prefix, start_idx, end_idx, len(group_cols), None))
                seen.add(prefix)

        # write merged headers per mapping
        for idx, (direction, start_idx, end_idx, num_cols, rid) in enumerate(mapping):
            start_letter = get_column_letter(start_idx)
            end_letter = get_column_letter(end_idx)
            ws.merge_cells(f"{start_letter}1:{end_letter}1")
            cell = ws.cell(row=1, column=start_idx, value=direction)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            fill = fills[idx % len(fills)]
            for col in range(start_idx, end_idx + 1):
                hcell = ws.cell(row=1, column=col)
                hcell.fill = fill; hcell.border = border; hcell.font = Font(bold=True)
                ws.column_dimensions[get_column_letter(col)].width = max(8, 90.0 / max(1, num_cols))

        # time column styling
        time_col_idx = cols.index("time") + 1 if "time" in cols else 1
        ws.cell(row=1, column=time_col_idx).fill = time_fill
        ws.cell(row=2, column=time_col_idx).fill = time_fill
        ws.column_dimensions[get_column_letter(time_col_idx)].width = 20

        # row2 header names
        for j, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=j, value=str(col_name))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if j == time_col_idx:
                cell.fill = time_fill
            else:
                for idx, (_, start_idx, end_idx, _, _) in enumerate(mapping):
                    if start_idx <= j <= end_idx:
                        cell.fill = fills[idx % len(fills)]; break
            cell.border = border

        # data rows
        combined_copy = combined.copy()
        for c in combined_copy.select_dtypes(include=["datetime", "datetimetz"]).columns:
            combined_copy[c] = pd.to_datetime(combined_copy[c]).dt.tz_localize(None)
        for i, row in enumerate(combined_copy.itertuples(index=False), start=3):
            for j, value in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.border = border
                if j == time_col_idx:
                    cell.fill = time_fill
                else:
                    for idx, (_, start_idx, end_idx, _, _) in enumerate(mapping):
                        if start_idx <= j <= end_idx:
                            cell.fill = fills[idx % len(fills)]; break
        last_row = 2 + combined_copy.shape[0]
        # outer border per group
        for (_, start_idx, end_idx, _, _) in mapping:
            for r in range(1, last_row + 1):
                for c in range(start_idx, end_idx + 1):
                    ws.cell(row=r, column=c).border = border

        return ws

    # Create a sheet per (version, divide_time) group
    summary_rows = [["Sheet Name", "Version", "Divide Time", "Records Included"]]
    for (version, divide_time), recs in groups.items():
        sheet_title = f"Auto_v{version}_d{divide_time}"
        created_ws = write_group_sheet(sheet_title, recs)
        if created_ws is not None:
            created_any = True
            summary_rows.append([created_ws.title, version, divide_time, ", ".join(str(rid) for rid, _, _ in recs)])

    # add summary sheet
    ws_summary = wb.create_sheet(title="Auto Summary")
    for r in summary_rows:
        ws_summary.append(r)
    for i in range(1, len(summary_rows[0]) + 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = 30

    # remove default sheet if unused
    if not created_any and default_sheet and default_sheet.title == "Sheet":
        # keep summary only
        pass
    elif default_sheet and default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    return wb, all_dfs_in_one


def compare_two_df(df1, df2, df1_name="Ground Truth", df2_name="To Check"):
    """
    df1 and df2 should have these columns: time, direction, count, portal
    Compare counts by (time, direction, portal). Return an openpyxl.Workbook
    with the same visual layout as the other exporters:
      - row1: merged direction headings
      - row2: portal/metric column names (e.g. "through abs", "through rel")
      - rows: time-aligned differences (abs and relative)
    """
    # helpers
    def get_portal_index(portal):
        portal_lower = str(portal).lower()
        if "through" in portal_lower:
            return 0
        elif "left" in portal_lower:
            return 1
        elif "right" in portal_lower:
            return 2
        else:
            return -1
    def get_portal_name(portal_index):
        return {0: "through", 1: "left", 2: "right"}.get(portal_index, "unknown")

    # defensive copies
    a = df1.copy()
    b = df2.copy()

    # normalize time column name if needed
    if "interval_start" in a.columns and "time" not in a.columns:
        a = a.rename(columns={"interval_start": "time"})
    if "interval_start" in b.columns and "time" not in b.columns:
        b = b.rename(columns={"interval_start": "time"})
    if "time" not in a.columns or "time" not in b.columns:
        raise ValueError("Both inputs must contain a 'time' (or 'interval_start') column.")

    a["time"] = pd.to_datetime(a["time"], errors="coerce")
    b["time"] = pd.to_datetime(b["time"], errors="coerce")

    # portal index / name
    a["portal_index"] = a["portal"].apply(get_portal_index)
    b["portal_index"] = b["portal"].apply(get_portal_index)
    a["portal_name"] = a["portal_index"].apply(get_portal_name)
    b["portal_name"] = b["portal_index"].apply(get_portal_name)

    # filter invalid portals
    a = a[a["portal_index"] != -1]
    b = b[b["portal_index"] != -1]

    # aggregate counts by time,direction,portal_name (in case of multiple rows)
    agg_a = a.groupby(["time", "direction", "portal_name"], as_index=False)["count"].sum().rename(columns={"count": f"count_{df1_name}"})
    agg_b = b.groupby(["time", "direction", "portal_name"], as_index=False)["count"].sum().rename(columns={"count": f"count_{df2_name}"})

    merged = pd.merge(agg_a, agg_b, on=["time", "direction", "portal_name"], how="outer")
    merged[f"count_{df1_name}"] = merged[f"count_{df1_name}"].fillna(0).astype(float)
    merged[f"count_{df2_name}"] = merged[f"count_{df2_name}"].fillna(0).astype(float)

    merged["abs_difference"] = (merged[f"count_{df1_name}"] - merged[f"count_{df2_name}"]).abs()
    # relative difference as abs / ground-truth (df1); None when ground truth is zero
    merged["relative_difference"] = merged.apply(
        lambda r: (r["abs_difference"] / r[f"count_{df1_name}"]) if r[f"count_{df1_name}"] != 0 else None,
        axis=1
    )

    if merged.empty:
        # return an empty workbook with a single sheet reporting no data
        wb_empty = openpyxl.Workbook()
        ws = wb_empty.active
        ws.title = "Comparison (empty)"
        ws.append(["No comparable data found"])
        return wb_empty

    # build wide dataframe: index = time, for each direction produce portal metric columns
    times = sorted(merged["time"].dropna().unique())
    wide = pd.DataFrame({"time": times})
    directions = merged["direction"].dropna().unique().tolist()

    for direction in directions:
        sub = merged[merged["direction"] == direction]
        portals = sub["portal_name"].dropna().unique().tolist()
        for portal in portals:
            s_abs = sub[sub["portal_name"] == portal][["time", "abs_difference"]].set_index("time").reindex(times)["abs_difference"]
            s_rel = sub[sub["portal_name"] == portal][["time", "relative_difference"]].set_index("time").reindex(times)["relative_difference"]
            col_abs = f"{direction} {portal} abs"
            col_rel = f"{direction} {portal} rel"
            wide[col_abs] = s_abs.values
            wide[col_rel] = s_rel.values

    # replace NaN with 0 for abs, keep None for relative where appropriate -> convert to floats
    wide = wide.sort_values("time").reset_index(drop=True)
    # convert timezone-aware datetimes to naive to be safe for Excel
    for c in wide.select_dtypes(include=["datetime", "datetimetz"]).columns:
        wide[c] = pd.to_datetime(wide[c]).dt.tz_localize(None)

    # create workbook and sheet with the same style as other exporters
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    from openpyxl.styles import PatternFill, Border, Side, Font
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    time_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fills = [
        PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    ]

    # prepare columns order: ensure time first
    cols = wide.columns.tolist()
    if cols[0] != "time":
        cols.remove("time")
        cols.insert(0, "time")

    # Build mapping for directions: determine for each direction how many columns it has
    mapping = []
    for idx, direction in enumerate(directions):
        # count how many cols belong to this direction
        dir_cols = [c for c in cols if c != "time" and str(c).startswith(f"{direction} ")]
        if not dir_cols:
            continue
        start_idx = cols.index(dir_cols[0]) + 1  # 1-based excel col index
        end_idx = cols.index(dir_cols[-1]) + 1
        mapping.append((direction, start_idx, end_idx, len(dir_cols)))

    # Row1: first cell empty then merged direction headings
    ws.cell(row=1, column=1, value="")
    for idx, (direction, start_idx, end_idx, num_cols) in enumerate(mapping):
        start_letter = get_column_letter(start_idx)
        end_letter = get_column_letter(end_idx)
        if start_idx <= end_idx:
            ws.merge_cells(f"{start_letter}1:{end_letter}1")
            hcell = ws.cell(row=1, column=start_idx, value=direction)
            hcell.alignment = Alignment(horizontal="center", vertical="center")
            fill = fills[idx % len(fills)]
            for col in range(start_idx, end_idx + 1):
                ccell = ws.cell(row=1, column=col)
                ccell.fill = fill
                ccell.border = border
                ccell.font = Font(bold=True)
                ws.column_dimensions[get_column_letter(col)].width = max(10, 90.0 / max(1, num_cols))

    # time column styling and width
    time_col_idx = cols.index("time") + 1
    ws.cell(row=1, column=time_col_idx).fill = time_fill
    ws.cell(row=2, column=time_col_idx).fill = time_fill
    ws.column_dimensions[get_column_letter(time_col_idx)].width = 20

    # Row2: actual column names
    for j, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=j, value=str(col_name))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if j == time_col_idx:
            cell.fill = time_fill
        else:
            for idx, (_, start_idx, end_idx, _) in enumerate(mapping):
                if start_idx <= j <= end_idx:
                    cell.fill = fills[idx % len(fills)]
                    break
        cell.border = border

    # Data rows starting at row3
    for i, row in enumerate(wide.itertuples(index=False), start=3):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = border
            if j == time_col_idx:
                cell.fill = time_fill
            else:
                for idx, (_, start_idx, end_idx, _) in enumerate(mapping):
                    if start_idx <= j <= end_idx:
                        cell.fill = fills[idx % len(fills)]
                        break

    # Outer border for each direction block
    last_data_row = 2 + wide.shape[0]
    for _, start_idx, end_idx, _ in mapping:
        for r in range(1, last_data_row + 1):
            for c in range(start_idx, end_idx + 1):
                ws.cell(row=r, column=c).border = border

    return wb
